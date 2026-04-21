import math, io, os, uuid
import cloudinary
import cloudinary.uploader
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import get_db
from app.core.security import get_current_user
from app.schemas.student import (
    StudentListResponse, StudentListItem, StudentCreate, StudentUpdate,
    ClubOut, StudentDetail, ImportResult, WeightVerifiedUpdate, StudentCardData,
)
from app.repositories import student_repo
from app.repositories.student_repo import REQUIRED_EXCEL_COLS
from app.repositories import tournament_repo
from typing import Optional
from app.core.constants import COMPETING_BELTS, WEIGHT_CLASSES
from app.core.config import settings

router = APIRouter(prefix="/students", tags=["students"])


async def _require_tournament_draft(db: AsyncSession = Depends(get_db)):
    """Raises 423 if the active tournament is not in DRAFT status."""
    status = await tournament_repo.get_active_tournament_status(db)
    if status is not None and status != "DRAFT":
        raise HTTPException(
            status_code=423,
            detail={
                "code": "TOURNAMENT_LOCKED",
                "message": f"Giải đấu đang ở trạng thái {status}. Không thể thêm/sửa/xóa vận động viên.",
            },
        )


@router.get("/meta")
async def get_meta(_=Depends(get_current_user)):
    return {"competing_belts": COMPETING_BELTS, "weight_classes": WEIGHT_CLASSES}


@router.get("/", response_model=StudentListResponse)
async def list_students(
    keyword         : Optional[str]   = Query(None),
    club_id         : Optional[int]   = Query(None),
    belt_rank       : Optional[str]   = Query(None),
    event           : Optional[str]   = Query(None),
    gender          : Optional[str]   = Query(None),
    dynamic_node_id : Optional[int]   = Query(None),
    weight_class    : Optional[str]   = Query(None),
    category_type   : Optional[str]   = Query(None),
    category_loai   : Optional[str]   = Query(None),
    quyen_selection : Optional[str]   = Query(None),
    weight_verified : Optional[bool]  = Query(None),
    tournament_id   : Optional[int]   = Query(None),
    status          : str             = Query("active"),
    page            : int             = Query(1, ge=1),
    page_size       : int             = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    # club role sees ALL students (no forced filter), edit/delete controlled on frontend+per-request
    rows, total, t_mode = await student_repo.get_students(db, keyword, club_id, belt_rank, event, gender, dynamic_node_id, weight_class, category_type, category_loai, quyen_selection, status, page, page_size, weight_verified, tournament_id)
    base_items = [dict(r) for r in rows]

    # For dynamic tournaments: enrich each item with registration display fields
    # node_id + node_name already available in rows from the JOIN in get_students
    reg_info: dict[int, dict] = {}
    if tournament_id and t_mode == "dynamic" and base_items:
        student_node_map = {
            d["id"]: {
                "node_id": d["swa_node_id"],
                "node_name": d["swa_node_name"],
                "category_name": d.get("swa_category_name"),
            }
            for d in base_items
            if d.get("swa_node_id") is not None
        }
        if student_node_map:
            reg_info = await student_repo.enrich_dynamic_registration(db, tournament_id, student_node_map)

    items = [StudentListItem(**{**d, **reg_info.get(d["id"], {})}) for d in base_items]
    return StudentListResponse(
        items=items, total=total, page=page, page_size=page_size,
        total_pages=max(1, math.ceil(total / page_size)),
    )


@router.get("/export-cards-data", response_model=list[StudentCardData])
async def export_cards_data(
    ids      : Optional[list[int]] = Query(None),
    club_id  : Optional[int]       = Query(None),
    tournament_id: Optional[int]   = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Chỉ Admin mới có thể export thẻ vận động viên"})
    if ids is None and club_id is None:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Cần cung cấp ids hoặc club_id"})
    if ids is not None and len(ids) == 0:
        raise HTTPException(status_code=422, detail={"code": "EMPTY_IDS", "message": "Danh sách ids không được rỗng"})
    rows = await student_repo.get_students_for_export(db, ids=ids, club_id=club_id, tournament_id=tournament_id)
    return [StudentCardData(**dict(r)) for r in rows]


@router.get("/clubs", response_model=list[ClubOut])
async def list_clubs(
    tournament_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    clubs = await student_repo.get_all_clubs(db, tournament_id=tournament_id)
    return [ClubOut(**dict(c)) for c in clubs]


@router.post("/", status_code=201)
async def create_student(
    body: StudentCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    if current_user.role in ("viewer", "referee"):
        raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Không có quyền tạo môn sinh"})
    # Club role: force club_id to own club
    if current_user.role == "club":
        if not current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Tài khoản club chưa được gán câu lạc bộ"})
        body.club_id = current_user.club_id
    if body.id_number and await student_repo.id_number_exists(db, body.id_number):
        raise HTTPException(400, detail={"code": "DUPLICATE_ID_NUMBER", "message": "CCCD đã tồn tại"})
    student = await student_repo.create_student(db, body.model_dump(), body.club_id)

    # Auto-register vào tournament LEGACY (dynamic tournaments bị bỏ qua bên trong hàm này).
    # Với dynamic tournament, dùng POST /tournaments/{id}/participants/{student_id}/register
    # hoặc POST /tournaments/{id}/participants/register-student (atomic).
    await student_repo.register_student_to_tournament(db, student.id)

    await db.commit()
    return {"id": student.id, "code": student.code, "message": "Tạo môn sinh thành công"}


@router.get("/{student_id}", response_model=StudentDetail)
async def get_student(
    student_id: int,
    tournament_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    data = await student_repo.get_student_detail(db, student_id, tournament_id=tournament_id)
    if not data:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Không tìm thấy môn sinh"})
    return data


@router.put("/{student_id}")
async def update_student(
    student_id: int,
    body: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    if current_user.role in ("viewer", "referee"):
        raise HTTPException(403, detail={"code": "FORBIDDEN"})
    if current_user.role == "club":
        if not current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN"})
        # Verify student belongs to this club
        student_club_id = await student_repo.get_student_club_id(db, student_id)
        if student_club_id != current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Bạn chỉ có thể sửa VĐV của câu lạc bộ mình"})
        body.club_id = current_user.club_id
    student = await student_repo.update_student(db, student_id, body.model_dump(exclude_none=True), body.club_id)
    if not student:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Không tìm thấy môn sinh"})

    # Sync legacy tournament registration nếu weight_class/category thay đổi.
    # Dynamic tournaments bị bỏ qua bên trong hàm này (structure_mode='dynamic').
    await student_repo.register_student_to_tournament(db, student_id)

    await db.commit()
    return {"id": student.id, "code": student.code, "message": "Cập nhật thành công"}


@router.patch("/{student_id}/weight-verified")
async def toggle_weight_verified(
    student_id: int,
    body: WeightVerifiedUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role in ("viewer", "referee"):
        raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Không có quyền cập nhật xác nhận hạng cân"})
    if current_user.role == "club":
        if not current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN"})
        student_club_id = await student_repo.get_student_club_id(db, student_id)
        if student_club_id != current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Bạn chỉ có thể cập nhật VĐV của câu lạc bộ mình"})
    from sqlalchemy import update as sa_update
    from app.models.student import Student
    result = await db.execute(
        sa_update(Student)
        .where(Student.id == student_id)
        .values(weight_verified=body.weight_verified)
        .returning(Student.id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Không tìm thấy môn sinh"})
    await db.commit()
    return {"id": student_id, "weight_verified": body.weight_verified}


@router.post("/import", response_model=ImportResult)
async def import_students(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    import openpyxl
    if current_user.role == "viewer":
        raise HTTPException(403, detail={"code": "FORBIDDEN"})
    if file.size and file.size > 10 * 1024 * 1024:
        raise HTTPException(400, detail={"code": "FILE_TOO_LARGE", "message": "File vượt quá 10MB"})
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_EXCEL_COLS - set(headers)
    if missing:
        raise HTTPException(400, detail={"code": "MISSING_COLUMNS", "message": f"Thiếu cột: {', '.join(missing)}"})
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    result = await student_repo.import_students_from_rows(db, rows)
    await db.commit()
    return result


@router.post("/{student_id}/avatar")
async def upload_avatar(
    student_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    if current_user.role in ("viewer", "referee"):
        raise HTTPException(403, detail={"code": "FORBIDDEN"})
    if current_user.role == "club":
        if not current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN"})
        student_club_id = await student_repo.get_student_club_id(db, student_id)
        if student_club_id != current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Bạn chỉ có thể cập nhật VĐV của câu lạc bộ mình"})
    allowed_types = {"image/jpeg", "image/png", "image/webp", "image/gif"}
    if file.content_type not in allowed_types:
        raise HTTPException(400, detail={"code": "INVALID_FILE_TYPE", "message": "Chỉ chấp nhận ảnh JPG, PNG, WEBP, GIF"})
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:
        raise HTTPException(400, detail={"code": "FILE_TOO_LARGE", "message": "Ảnh không được vượt quá 2MB"})

    # Delete old avatar from Cloudinary if present
    old_url = await student_repo.get_student_avatar(db, student_id)
    if old_url and "cloudinary.com" in old_url:
        # public_id is the path segment after /upload/vX/
        try:
            parts = old_url.split("/upload/")
            if len(parts) == 2:
                public_id = parts[1].rsplit(".", 1)[0]  # strip extension
                # strip version prefix e.g. "v1234567890/"
                if "/" in public_id:
                    public_id = public_id.split("/", 1)[1]
                cloudinary.uploader.destroy(public_id)
        except Exception:
            pass  # non-critical, continue upload

    result = cloudinary.uploader.upload(
        io.BytesIO(contents),
        folder="vovinam/avatars",
        resource_type="image",
    )
    avatar_url = result["secure_url"]
    await student_repo.update_avatar_url(db, student_id, avatar_url)
    await db.commit()
    return {"avatar_url": avatar_url}


@router.delete("/bulk")
async def bulk_delete_students(
    ids: list[int],
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    if current_user.role != "admin":
        raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Chỉ Admin mới có thể xóa hàng loạt"})
    if not ids:
        raise HTTPException(400, detail={"code": "BAD_REQUEST", "message": "Danh sách ids không được rỗng"})
    count = await student_repo.bulk_soft_delete_students(db, ids)
    await db.commit()
    return {"deleted": count, "message": f"Đã xoá {count} môn sinh"}


@router.delete("/{student_id}")
async def delete_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _lock=Depends(_require_tournament_draft),
):
    if current_user.role in ("viewer", "referee"):
        raise HTTPException(403, detail={"code": "FORBIDDEN"})
    if current_user.role == "club":
        if not current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN"})
        student_club_id = await student_repo.get_student_club_id(db, student_id)
        if student_club_id != current_user.club_id:
            raise HTTPException(403, detail={"code": "FORBIDDEN", "message": "Bạn chỉ có thể xóa VĐV của câu lạc bộ mình"})
    ok = await student_repo.soft_delete_student(db, student_id)
    if not ok:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Không tìm thấy môn sinh"})
    await db.commit()
    return {"message": "Đã xoá môn sinh"}
